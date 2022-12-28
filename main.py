import csv
import itertools
import re
from enum import Enum
from typing import List, Dict

import matplotlib.pyplot as plt
import numpy
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.borders import BORDER_THIN as thin
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00 as percentage
from jinja2 import Environment, FileSystemLoader
import pdfkit


class DataSet:
    def __init__(self, file: str, job: str):
        self.file_name = file
        self.vacancies_objects = [Vacancy(vac) for vac in self.csv_reader(self.file_name)]

    @staticmethod
    def clean_html(raw_html: str) -> str:
        return ' '.join(re.sub(re.compile('<.*?>'), '', raw_html).replace('\n', ';').split())

    @staticmethod
    def csv_reader(file: str) -> (List[str], List[List[str]]):
        with open(file, 'r', encoding="utf-8-sig") as csvfile:
            csvreader = csv.reader(csvfile)
            data = []
            titles = csvreader.__next__()
            inputconnect = InputConect()
            for vac in csvreader:
                dic = {}
                vac = list(filter(lambda x: x != '', vac))
                if len(vac) == len(titles):
                    for i in range(len(vac)):
                        dic[titles[i]] = vac[i]
                    data.append(dic)
                    inputconnect.count(Vacancy(dic), job)
            rep = Report
            inputconnect.print()
            rep.generate_image(job=job, data=inputconnect)
            rep.generate_excel(job=job, data=inputconnect)
            rep.generate_pdf(job=job,data=inputconnect)
            return data


class Salary:
    def __init__(self, salary_from, salary_to, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency


class Vacancy:
    def __init__(self, dic_vac):
        self.name = dic_vac['name']
        self.salary = Salary(dic_vac['salary_from'], dic_vac['salary_to'], dic_vac['salary_currency'])
        self.area_name = dic_vac['area_name']
        self.published_at = dic_vac['published_at']

    def __iter__(self):
        self.row = {'name': self.name,
                    'salary_from': self.salary,
                    'area_name': self.area_name,
                    'published_at': self.published_at, }
        return iter(self.row)

    def __getitem__(self, item):
        self.row = {'name': self.name,
                    'salary_from': self.salary,
                    'area_name': self.area_name,
                    'published_at': self.published_at, }
        return self.row[item]


class InputConect:
    def __init__(self):
        self.years_sal_all = {}
        self.years_count_all = {}
        self.years_sal_job = {}
        self.years_count_job = {}
        self.city_sal = {}
        self.city_percent = {}

    def count(self, vac: Vacancy, job: str):
        self.years_sal_all = self.years_info_sal_all(vac, self.years_sal_all)
        self.years_count_all = self.years_info_count_all(vac, self.years_count_all)
        self.years_sal_job = self.years_info_sal_job(vac, job, self.years_sal_job)
        self.years_count_job = self.years_info_count_job(vac, job, self.years_count_job)
        self.city_sal = self.city_info_sal(vac, self.city_sal)
        self.city_percent = self.city_info_percent(vac, self.city_percent)

    def print(self):
        self.sorting()
        print('Динамика уровня зарплат по годам: ' + str(self.years_sal_all))
        print('Динамика количества вакансий по годам: ' + str(self.years_count_all))
        print('Динамика уровня зарплат по годам для выбранной профессии: ' + str(self.years_sal_job))
        print('Динамика количества вакансий по годам для выбранной профессии: ' + str(self.years_count_job))
        print('Уровень зарплат по городам (в порядке убывания): ' + str(self.city_sal))
        print('Доля вакансий по городам (в порядке убывания): ' + str(self.city_percent))

    def sorting(self):
        self.years_sal_all = self.get_avg_val(self.years_sal_all, self.years_count_all)
        self.years_sal_job = self.get_avg_val(self.years_sal_job, self.years_count_job)
        self.city_sal = self.get_avg_val(self.city_sal, self.city_percent)
        total_cities = 0
        for key, value in self.city_percent.items():
            total_cities += value
        self.city_percent = self.get_avg_count(self.city_percent, total_cities)
        for key, value in self.city_sal.copy().items():
            if not self.city_percent.keys().__contains__(key):
                self.city_sal.pop(key)
        self.city_sal = dict(sorted(self.city_sal.items(), key=lambda x: x[1], reverse=True))
        self.city_percent = dict(sorted(self.city_percent.items(), key=lambda x: x[1], reverse=True))
        self.city_sal = dict(itertools.islice(self.city_sal.items(), 10))
        self.city_percent = dict(itertools.islice(self.city_percent.items(), 10))

    def years_info_sal_all(self, vac: Vacancy, years_sal: Dict):
        if years_sal.keys().__contains__(self.get_correct_data(vac.published_at)):
            years_sal[self.get_correct_data(vac.published_at)] += self.sort_money(vac)
        else:
            years_sal[self.get_correct_data(vac.published_at)] = self.sort_money(vac)
        return years_sal

    def years_info_sal_job(self, vac: Vacancy, job: str, years_sal: Dict):
        if job in vac.name:
            if years_sal.keys().__contains__(self.get_correct_data(vac.published_at)):
                years_sal[self.get_correct_data(vac.published_at)] += self.sort_money(vac)
            else:
                years_sal[self.get_correct_data(vac.published_at)] = self.sort_money(vac)
        else:
            if not years_sal.keys().__contains__(self.get_correct_data(vac.published_at)):
                years_sal[self.get_correct_data(vac.published_at)] = 0
        return years_sal

    def years_info_count_all(self, vac: Vacancy, years_count: Dict):
        if years_count.keys().__contains__(self.get_correct_data(vac.published_at)):
            years_count[self.get_correct_data(vac.published_at)] += 1
        else:
            years_count[self.get_correct_data(vac.published_at)] = 1
        return years_count

    def years_info_count_job(self, vac: Vacancy, job: str, years_count: Dict):
        if job in vac.name:
            if years_count.keys().__contains__(self.get_correct_data(vac.published_at)):
                years_count[self.get_correct_data(vac.published_at)] += 1
            else:
                years_count[self.get_correct_data(vac.published_at)] = 1
        else:
            if not years_count.keys().__contains__(self.get_correct_data(vac.published_at)):
                years_count[self.get_correct_data(vac.published_at)] = 0
        return years_count

    def city_info_sal(self, vac: Vacancy, city_sal):
        if city_sal.keys().__contains__(vac.area_name):
            city_sal[vac.area_name] += self.sort_money(vac)
        else:
            city_sal[vac.area_name] = self.sort_money(vac)
        return city_sal

    def city_info_percent(self, vac: Vacancy, city_percent: Dict):
        if city_percent.keys().__contains__(vac.area_name):
            city_percent[vac.area_name] += 1
        else:
            city_percent[vac.area_name] = 1
        return city_percent

    def get_avg_val(self, dic_sal: Dict, dic_count: Dict):
        for key in dic_sal:
            if dic_count[key] != 0:
                dic_sal[key] = int(dic_sal[key] / dic_count[key])
        return dic_sal

    def get_avg_count(self, dic_sal: Dict, count: int):
        for key in dic_sal.copy():
            dic_sal[key] = dic_sal[key] / count
            if dic_sal[key] <= 0.01:
                dic_sal.pop(key)
            if dic_sal.__contains__(key):
                dic_sal[key] = round(dic_sal[key], 4)
        return dic_sal

    def get_correct_data(self, date: str):
        return int(date[0:4])

    def sort_money(self, vac: Vacancy):
        salary_from = int(vac.salary.salary_from.split('.')[0])
        salary_to = int(vac.salary.salary_to.split('.')[0])
        return (salary_from + salary_to) / 2 * currency_to_rub[vac.salary.salary_currency]


class Report:
    @staticmethod
    def generate_excel(job: str, data: InputConect):
        wb = Workbook()
        ws_year = wb.active
        ws_year.title = "Статистика по годам"
        columns_year = ['Год', 'Средняя зарплата', f'Средняя зарплата по професии - {job}',
                        'Количество вакансий', f'Количество вакансий {job} в год']
        Report.name_column(columns_year, ws_year)
        for key, value in data.years_sal_all.items():
            ws_year.append([key, value, data.years_sal_job[key], data.years_count_all[key], data.years_count_job[key]])
        Report.borders_width(ws_year)
        ws_city = wb.create_sheet("Статистика по городам")
        columns_city = ['Город', 'Уровень Зарплат', 'Доля вакансий']
        Report.name_column(columns_city, ws_city)
        for key, value in data.city_sal.items():
            ws_city.append([key, value, data.city_percent[key]])
        for i in range(2, 12):
            ws_city[f'C{i}'].number_format = percentage
        Report.borders_width(ws_city)
        wb.save(filename="report.xlsx")

    @staticmethod
    def name_column(columns, ws):
        for i, column in enumerate(columns):
            ws.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)

    @staticmethod
    def borders_width(ws):
        for column in ws.columns:
            length = max(len(str(cell.value)) for cell in column)
            for cell in column:
                cell.border = Border(left=Side(thin), top=Side(thin), right=Side(thin), bottom=Side(thin))
            ws.column_dimensions[column[0].column_letter].width = length + 1

    @staticmethod
    def generate_image(job: str, data: InputConect):
        figure, ((sal_graph, сount_graph), (city_sal, city_job)) = plt.subplots(2, 2)
        width = 0.4
        x_axis_years = np.arange(len(data.years_sal_all.keys()))

        Report.create_diagramm_sal(data, job, sal_graph, width, x_axis_years, 'Уровень зарплат по годам')
        Report.create_diagramm_vacs(data, job, width, x_axis_years, сount_graph)
        Report.create_invert_diagramm(city_sal, data, job)
        Report.create_pie_charm(city_job, data, job)

        plt.tight_layout()
        plt.savefig('graph.png')

    @staticmethod
    def create_diagramm_vacs(data, job, width, x_axis_years, сount_graph):
        сount_graph.set_title('Количество вакансий по год')
        сount_graph.legend(fontsize=8)
        сount_graph.bar(x_axis_years - width / 2, data.years_count_all.values(), width=width,
                        label='Количество вакансий в год')
        сount_graph.bar(x_axis_years + width / 2, data.years_count_job.values(), width=width,
                        label=f'Количество вакансий в год для {job}')
        сount_graph.set_xticks(x_axis_years, data.years_sal_job.keys(), rotation='vertical')
        сount_graph.tick_params(axis='both', labelsize=8)
        сount_graph.grid(True, axis='y')

    @staticmethod
    def create_diagramm_sal(data, job, sal_graph, width, x_axis_years, text):
        sal_graph.set_title(f'{text}')
        sal_graph.legend(fontsize=8)
        sal_graph.bar(x_axis_years - width / 2, data.years_sal_all.values(), width=width,
                      label=f'Средняя з/п в год')
        sal_graph.bar(x_axis_years + width / 2, data.years_sal_job.values(), width=width,
                      label=f'Средняя з/п в год для {job}')
        sal_graph.set_xticks(x_axis_years, data.years_sal_job.keys(), rotation='vertical')
        sal_graph.tick_params(axis='both', labelsize=8)
        sal_graph.grid(True, axis='y')

    @staticmethod
    def create_invert_diagramm(city_sal, data, job):
        city_sal.invert_yaxis()
        y_axis_cities = list(data.city_sal.keys())
        city_sal.barh(y_axis_cities, data.city_sal.values())
        city_sal.set_yticklabels(y_axis_cities, fontsize=6, va='center', ha='right')
        city_sal.set_title(job)
        city_sal.tick_params(axis='both', labelsize=8)
        city_sal.grid(True, axis='x')

    @staticmethod
    def create_pie_charm(city_job, data, job):
        value = data.city_percent
        other = 1 - sum((list(value.values())))
        other_dic = {'Другие': other}
        other_dic.update(value)
        city_job.set_title(job)
        city_job.pie(list(other_dic.values()), labels=list(other_dic.keys()), textprops={'fontsize': 6})
        city_job.axis('scaled')

    @staticmethod
    def generate_pdf(job:str, data:InputConect):
        image="graph.png"
        year_headers = ['Год', 'Средняя зарплата', f'Средняя зарплата по професии - {job}',
                        'Количество вакансий', f'Количество вакансий {job} в год']
        city_headers = ['Город', 'Уровень Зарплат', 'Доля вакансий']
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        year_data = {year: [salary, salary_job, count, count_job]
                     for year, salary, salary_job, count, count_job in zip(data.years_sal_all.keys(),
                                                                           data.years_sal_all.values(),
                                                                           data.years_sal_job.values(),
                                                                           data.years_count_all.values(),
                                                                           data.years_count_job.values())
                     }
        city_data = {city: [salary, ratio]
                     for city, salary, ratio in zip(data.city_sal.keys(),
                                                    data.city_sal.values(),
                                                    data.city_percent.values())
                     }
        pdf_templ = template.render( {'image_file': image,
             'image_style': 'style="max-width:1024px; max-height:680px"',
             'salary_data': year_data,
             'city_data': city_data,
             'header_year': year_headers,
             'header_city': city_headers,
             'profession_name': f"{job}",
             'h1_style': 'style="text-align:center; font-size:32px"',
             'h2_style': 'style="text-align:center"',
             'cell_style_none': "style=''",
             'cell_style': 'style="border:1px solid black; border-collapse: collapse; font-size: 16px; height: 19pt;'
                           'padding: 5px; text-align:center"'})
        config = pdfkit.configuration(wkhtmltopdf=r'E:\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_templ, job, configuration=config, options={'enable-local-file-access': None})


class Translate(Enum):
    name = 'Название'
    description = 'Описание'
    key_skills = 'Навыки'
    experience_id = 'Опыт работы'
    premium = 'Премиум-вакансия'
    employer_name = 'Компания'
    salary_from = 'Оклад'
    salary_to = 'Верхняя граница вилки оклада'
    salary_gross = 'Оклад указан до вычета налогов'
    salary_currency = 'Идентификатор валюты оклада'
    area_name = 'Название региона'
    published_at = 'Дата публикации вакансии'
    AZN = "Манаты"
    BYR = "Белорусские рубли"
    EUR = "Евро"
    GEL = "Грузинский лари"
    KGS = "Киргизский сом"
    KZT = "Тенге"
    RUR = "Рубли"
    UAH = "Гривны"
    USD = "Доллары"
    UZS = "Узбекский сум"
    noExperience = "Нет опыта"
    between1And3 = "От 1 года до 3 лет"
    between3And6 = "От 3 до 6 лет"
    moreThan6 = "Более 6 лет"


currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

translate_yes_no = {
    'да': True,
    'нет': False,
    'true': 'Да',
    'false': 'Нет',
}

file_name = input('Введите название файла: ')
job = input('Введите название профессии: ')
dataset = DataSet(file_name, job)
